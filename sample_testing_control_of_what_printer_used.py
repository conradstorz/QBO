"""This is ChatGPT code. It was created on 1/30/2024
It has been tested and is partially working.
"""

import win32print
import win32ui

import json
# Custom function to serialize the dictionary excluding unserializable objects
def custom_json_serializer(obj):
    # usage example: pretty_dict = json.dumps(dict, default=custom_json_serializer, indent=4)
    if isinstance(obj, (int, float, str, bool, list, tuple, dict, type(None))):
        return obj
    else:
        return str(obj)
    
from openpyxl import load_workbook
from pathlib import Path
from loguru import logger

""" ---future---
# import my custom code
from custom_loguru import defineLoggers
from time_strings import LOCAL_NOW_STRING

# gather runtime info
RUNTIME_NAME = Path(__file__).name
RUNTIME_CWD = Path.cwd()
LOGNAME = f"{RUNTIME_NAME}"
print(f"Logging file: {LOGNAME}")

# setup logging
defineLoggers(LOGNAME)
"""

# Define the path to your Excel spreadsheet
spreadsheet_path = 'Outputfile0.xlsx'

# Define the name of the printer you want to use
printer_name = 'Canon TR8500 series'  # Replace with your printer's name

# Load the Excel spreadsheet
workbook = load_workbook(spreadsheet_path)

# List available sheet names in the workbook
sheet_names = workbook.sheetnames

# Choose the first sheet (index 0)
worksheet = workbook[sheet_names[0]]

# Create a Windows printer handle
printer_handle = win32print.OpenPrinter(printer_name)

# Set up the printer to print the worksheet
printer_info = win32print.GetPrinter(printer_handle, 2)

# Pretty print and log the dictionary item
pretty_json = json.dumps(printer_info, default=custom_json_serializer, indent=4)
logger.info(pretty_json)

# Create a printer device context
printer_dc = win32ui.CreateDC()
printer_dc.CreatePrinterDC(printer_name)

# Set up the printing job
printer_dc.StartDoc('Your_Print_Document_Name')
printer_dc.StartPage()

# Print a sample text
sample_text = "Hello, World!"
printer_dc.TextOut(100, 100, sample_text)

# Print the worksheet
# You may need to adjust the positioning and size of the printed content
# based on your specific requirements
try:
    for row in worksheet.iter_rows():
        for cell in row:
            logger.debug(f"Cell Value: {cell.value}")
            logger.debug(f"Cell Coordinates: ({cell.column}, {cell.row})")
            printer_dc.TextOut(cell.column, cell.row, str(cell.value))
except Exception as e:
    print(f"Printing Error: {e}")


# End the printing job
printer_dc.EndPage()
printer_dc.EndDoc()

# Close the printer handle
win32print.ClosePrinter(printer_handle)
